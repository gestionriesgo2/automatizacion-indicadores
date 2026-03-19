import pandas as pd
from openpyxl import load_workbook
import re


def procesar_fichas_drive(
    files_anio,
    banco,
    col_codigo,
    read_excel_from_drive,
    clean_str,
    norm_code
):
    """
    Procesa las fichas del año y actualiza el banco.

    Retorna:
        banco actualizado
        registros (log)
    """

    registros = []
    
    if "RANGO DE GESTION" not in banco.columns:
        banco["RANGO DE GESTION"] = None

    for f in files_anio:
        filename = f["name"]
        file_id = f["id"]

        if not filename.lower().endswith((".xlsx", ".xlsm")):
            continue

        try:
            stream = read_excel_from_drive(file_id)
            wb = load_workbook(stream, data_only=True, read_only=True)
            nombre_base = filename.lower().replace(".xlsx", "").replace(".xlsm", "")

            # ----------------------------
            # Buscar hoja principal
            # ----------------------------
            posibles_codigos = [p for p in nombre_base.split() if "ind" in p and "-" in p]
            nombre_hoja = None

            for hoja in wb.sheetnames:
                hoja_l = hoja.lower()
                if any(c in hoja_l for c in posibles_codigos):
                    nombre_hoja = hoja
                    break
                if nombre_base in hoja_l:
                    nombre_hoja = hoja
                    break

            if not nombre_hoja:
                nombre_hoja = next(
                    (h for h in wb.sheetnames if "ficha" in h.lower() or "indicador" in h.lower()),
                    wb.sheetnames[0]
                )

            ws = wb[nombre_hoja]

            # ----------------------------
            # Extraer código
            # ----------------------------
            # Obtener código directamente desde la celda correcta
            codigo = ws["L5"].value or ws["M5"].value

            # Normalizar si existe valor
            if codigo:
                codigo = norm_code(codigo)

            # Validar que realmente sea un código válido
            if not codigo or not str(codigo).startswith("IND-"):
                codigo = None


            # ----------------------------
            # Validar código
            # ----------------------------
            if not codigo:
                registros.append({
                    "archivo": filename,
                    "hoja": nombre_hoja,
                    "codigo": None,
                    "accion": "codigo_no_encontrado",
                    "ok": False
                })
                continue


            # ----------------------------
            # Datos principales
            # ----------------------------
            fila = {
                col_codigo: codigo,
                "INDICADOR": clean_str(ws["C5"].value),
                "JERARQUÍA": clean_str(ws["I5"].value),
                "PROCESO": clean_str(ws["H7"].value),
                "OBJETIVO-DESCRIPCIÓN": clean_str(ws["C6"].value),
                "ÁREA": clean_str(ws["C7"].value),
                "TIPO DE INDICADOR": clean_str(ws["C8"].value),
                "TENDENCIA": clean_str(ws["L8"].value),
                "FÓRMULA": f"{clean_str(ws['C9'].value)} / {clean_str(ws['H9'].value)}",
                "FUENTE NUMERADOR": clean_str(ws["C10"].value),
                "FUENTE DENOMINADOR": clean_str(ws["H10"].value),
                "PERIODICIDAD MEDICION": clean_str(ws["C11"].value),
                "PERIODICIDAD ANÁLISIS": clean_str(ws["C12"].value),
                "OBSERVACIONES": clean_str(ws["C13"].value),
                "NORMA RELACIONADA": clean_str(ws["L9"].value),
                "Critico": clean_str(ws["K11"].value),
                "Aceptable": clean_str(ws["L11"].value),
                "Satisfactorio": clean_str(ws["M11"].value),
                "VALORACIÓN": clean_str(ws["O19"].value),
                "RANGO DE GESTION": clean_str(ws["P19"].value)
            }

            # ----------------------------
            # Meses
            # ----------------------------
            mapa_meses = {
                "ene-25": "B19", "feb-25": "C19", "mar-25": "D19", "abr-25": "E19",
                "may-25": "F19", "jun-25": "G19", "jul-25": "H19", "ago-25": "I19",
                "sept-25": "J19", "oct-25": "K19", "nov-25": "L19", "dic-25": "M19"
            }

            valores = []
            valores_limpios = []
            es_porcentaje = False

            for mes, celda in mapa_meses.items():
                v = ws[celda].value

                if v is None or str(v).strip() == "" or "N/A" in str(v) or "#" in str(v):
                    fila[mes] = "N/A"
                    valores.append(None)
                    continue

                try:
                    # Si viene como texto con %
                    if isinstance(v, str) and "%" in v:
                        num = float(v.replace("%", "").replace(",", "."))
                        es_porcentaje = True
                    else:
                        num = float(v)

                        # Si viene como decimal tipo 0.85 → es porcentaje Excel
                        if 0 < num <= 1:
                            num = num * 100
                            es_porcentaje = True

                    valores.append(num)
                    valores_limpios.append(num)

                    # Guardar formateado respetando tipo
                    if es_porcentaje:
                        fila[mes] = f"{num:.2f}%"
                    else:
                        fila[mes] = round(num, 2)

                except:
                    fila[mes] = "N/A"
                    valores.append(None)

            # ----------------------------
            # VALOR ANUAL desde Excel (N19)
            # ----------------------------
            v_anual = ws["N19"].value

            if v_anual is None or str(v_anual).strip() == "" or "N/A" in str(v_anual) or "#" in str(v_anual):
                fila["VALOR ANUAL"] = ""
            else:
                try:
                    if isinstance(v_anual, str) and "%" in v_anual:
                        num = float(v_anual.replace("%", "").replace(",", "."))
                        fila["VALOR ANUAL"] = f"{round(num)}%"

                    else:
                        num = float(v_anual)

                        if es_porcentaje:
                            if 0 < num <= 1:
                                num = num * 100

                            fila["VALOR ANUAL"] = f"{round(num)}%"
                        else:
                            fila["VALOR ANUAL"] = round(num, 2)

                except:
                    fila["VALOR ANUAL"] = ""


            # ----------------------------
            # Hoja evaluación
            # ----------------------------
            hoja_eval = next((wb[h] for h in wb.sheetnames if "eval" in h.lower()), None)
            if hoja_eval:
                fila.update({
                    "ESTADO DEL INDICADOR": clean_str(hoja_eval["A2"].value),
                    "ORIGEN": clean_str(hoja_eval["B2"].value),
                    "DOCUMENTADO": clean_str(hoja_eval["C2"].value),
                    "DE SEG CONTRACTUAL": clean_str(hoja_eval["D2"].value),
                    "REVISADOS": clean_str(hoja_eval["E2"].value),
                })

            # Mantener solo columnas válidas
            fila = {k: v for k, v in fila.items() if k in banco.columns}

            # ----------------------------
            # Actualizar banco
            # ----------------------------
            if codigo in banco[col_codigo].values:
                for k, v in fila.items():
                    banco.loc[banco[col_codigo] == codigo, k] = v
                accion = "actualizado"
            else:
                nueva = {c: None for c in banco.columns}
                nueva.update(fila)
                banco = pd.concat([banco, pd.DataFrame([nueva])], ignore_index=True)
                accion = "agregado"

            registros.append({
                "archivo": filename,
                "hoja": nombre_hoja,
                "codigo": codigo,
                "accion": accion,
                "ok": True
            })

        except Exception as e:
            registros.append({
                "archivo": filename,
                "hoja": None,
                "codigo": None,
                "accion": f"error: {e}",
                "ok": False
            })
    # ============================
    # ORDENAR BANCO
    # ============================


    def extraer_numero_codigo(codigo):
        if pd.isna(codigo):
            return 0
        match = re.search(r'(\d+)$', str(codigo))
        return int(match.group(1)) if match else 0

    banco["_orden"] = banco[col_codigo].apply(extraer_numero_codigo)

    banco = banco.sort_values(by=["ÁREA", "_orden"])

    banco = banco.drop(columns=["_orden"])

    banco = banco.reset_index(drop=True)

    return banco, registros

