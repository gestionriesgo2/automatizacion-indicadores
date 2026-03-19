import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from components.resumen import generar_resumenes
from components.modelo_atencion import generar_resumen_modelo_atencion

import pandas as pd


def guardar_banco_con_estilos_drive(
    banco,
    create_or_update_file,
    banco_file_id,
    banco_folder_id,
    filename="Banco_Indicadores.xlsx"
):
    """
    Aplica estilos al banco y lo guarda en Google Drive.
    """

    print("🎨 Aplicando estilos y guardando Banco en Drive...")

    # --------------------------------------------------
    # 🔹 Normalizar columnas del banco (MUY IMPORTANTE)
    # --------------------------------------------------
    banco.columns = banco.columns.str.strip().str.upper()

    # --------------------------------------------------
    # 🔹 Generar resúmenes
    # --------------------------------------------------
    resumenes = generar_resumenes(banco)
    modelo_atencion = generar_resumen_modelo_atencion(banco)

    # 🔎 Debug opcional (puedes quitarlo luego)
    print("Claves disponibles en resumenes:", resumenes.keys())

    # --------------------------------------------------
    # 🔹 Obtener dataframes de forma segura
    # --------------------------------------------------
    resumen_area = resumenes["resumen_area"]
    resumen_estado = resumenes["resumen_estado"]
    resumen_periodicidad = resumenes["resumen_periodicidad"]
    resumen_general = resumenes["resumen_general"]
    resumen_jerarquia = resumenes["resumen_jerarquia"]
    resumen_tipo = resumenes["resumen_tipo"]
    resumen_cumple = resumenes["resumen_cumple"]

    # Validación por si falta alguna clave
    for nombre, df in {
        "area": resumen_area,
        "estado": resumen_estado,
        "periodicidad": resumen_periodicidad,
        "general": resumen_general,
        "cumple": resumen_cumple,
        "jerarquia": resumen_jerarquia,
        "tipo": resumen_tipo
    }.items():
        if df is None:
            raise ValueError(f"❌ No se encontró la clave '{nombre}' en generar_resumenes()")

    # --------------------------------------------------
    # Guardar banco en memoria
    # --------------------------------------------------
    buffer = io.BytesIO()
    banco.to_excel(buffer, index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    # --------------------------------------------------
    # Estilos
    # --------------------------------------------------
    header_fill = PatternFill(start_color="A7D08C", end_color="A7D08C", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    columnas_sin_color = ["Q", "R", "S"]

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=1, column=col)
        if col_letter not in columnas_sin_color:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    ws["R1"].fill = PatternFill(start_color="FF0000", fill_type="solid")
    ws["S1"].fill = PatternFill(start_color="FFFF00", fill_type="solid")
    ws["T1"].fill = PatternFill(start_color="00FF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 30

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "C2"

    # ==================================================
    # 📊 AGREGAR HOJAS DE RESUMEN
    # ==================================================

    def agregar_hoja_resumen(nombre, df):
        if df is None or df.empty:
            print(f"⚠️ Hoja {nombre} vacía, no se crea.")
            return

        ws_res = wb.create_sheet(title=nombre)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_res.cell(row=r_idx, column=c_idx, value=value)

        for col in range(1, ws_res.max_column + 1):
            cell = ws_res.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for row in ws_res.iter_rows(min_row=2, max_row=ws_res.max_row, max_col=ws_res.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for col in ws_res.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            col_letter = get_column_letter(col[0].column)
            ws_res.column_dimensions[col_letter].width = min(max_len + 2, 40)

        ws_res.auto_filter.ref = f"A1:{get_column_letter(ws_res.max_column)}{ws_res.max_row}"
        ws_res.freeze_panes = "A2"

    # Crear hojas
    agregar_hoja_resumen("Resumen_Area", resumen_area)
    agregar_hoja_resumen("Resumen_Estado", resumen_estado)
    agregar_hoja_resumen("Resumen_Periodicidad", resumen_periodicidad)
    agregar_hoja_resumen("Resumen_General", resumen_general)
    agregar_hoja_resumen("Resumen_Jerarquia", resumen_jerarquia)
    agregar_hoja_resumen("Resumen_Tipo", resumen_tipo)
    agregar_hoja_resumen("Resumen_Cumple", resumen_cumple)
    agregar_hoja_resumen("Modelo_Atencion", modelo_atencion)
    
    
    # ==================================================
    # 🎯 FORMATO ESPECIAL HOJA MODELO_ATENCION
    # ==================================================

    ws_modelo = wb["Modelo_Atencion"]
    
    # ==================================================
    # 🎯 AJUSTAR FILTRO PARA QUE ESTÉ EN FILA 2
    # ==================================================

    max_col_letter = get_column_letter(ws_modelo.max_column)
    max_row = ws_modelo.max_row

    # Quitar filtro anterior (el que se puso en fila 1)
    ws_modelo.auto_filter.ref = None

    # Aplicar filtro desde fila 2
    ws_modelo.auto_filter.ref = f"A2:{max_col_letter}{max_row}"

    # Congelar hasta fila 3 (para mantener visibles las 2 filas de encabezado)
    ws_modelo.freeze_panes = "A3"

    # Detectar últimas 3 filas (VALORACIÓN GENERAL)
    fila_inicio = ws_modelo.max_row - 2
    f1 = fila_inicio
    f2 = fila_inicio + 1
    f3 = fila_inicio + 2

    # --------------------------------------------------
    # 1️⃣ Combinar A:D para VALORACIÓN GENERAL
    # --------------------------------------------------
    ws_modelo.merge_cells(f"A{f1}:D{f3}")

    celda = ws_modelo[f"A{f1}"]
    celda.value = "VALORACIÓN GENERAL"
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.font = Font(bold=True)

    # --------------------------------------------------
    # 2️⃣ Combinar Meta + Medición por trimestre
    # (según estructura actual)
    # --------------------------------------------------

    # T1 → columnas E y F
    ws_modelo.merge_cells(f"E{f1}:F{f1}")
    ws_modelo.merge_cells(f"E{f2}:F{f2}")
    ws_modelo.merge_cells(f"E{f3}:F{f3}")

    # T2 → columnas H y I
    ws_modelo.merge_cells(f"H{f1}:I{f1}")
    ws_modelo.merge_cells(f"H{f2}:I{f2}")
    ws_modelo.merge_cells(f"H{f3}:I{f3}")

    # T3 → columnas K y L
    ws_modelo.merge_cells(f"K{f1}:L{f1}")
    ws_modelo.merge_cells(f"K{f2}:L{f2}")
    ws_modelo.merge_cells(f"K{f3}:L{f3}")

    # T4 → columnas N y O
    ws_modelo.merge_cells(f"N{f1}:O{f1}")
    ws_modelo.merge_cells(f"N{f2}:O{f2}")
    ws_modelo.merge_cells(f"N{f3}:O{f3}")

    # --------------------------------------------------
    # Subir a Drive
    # --------------------------------------------------
    buffer_out = io.BytesIO()
    wb.save(buffer_out)
    buffer_out.seek(0)

    create_or_update_file(
        bytes_data=buffer_out.getvalue(),
        file_id=banco_file_id,
        filename=filename,
        parent_folder_id=banco_folder_id
    )

    print("✅ Banco guardado correctamente en Drive")