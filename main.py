from drive_reader import list_files_in_folder, read_excel_from_drive, upload_bytes_to_drive
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------
# ARCHIVOS EN GOOGLE DRIVE
# --------------------------------
BANCO_FILE_ID = "1m6MJt-aGnAlvAhy0WzZzI-VM92oCaA5o"
FICHAS_FOLDER_ID = "1dpc-JNrZbww9ud8O-VQQ5EqS9VyauIiP"

# ----------------------------
# FUNCIONES AUXILIARES
# ----------------------------
def clean_str(x):
    if x is None:
        return None
    s = str(x).strip()
    return s if s != "" else None

def norm_code(x):
    if x is None:
        return None
    return str(x).strip().upper().replace(" ", "")

# ----------------------------
# CARGAR BANCO DESDE DRIVE
# ----------------------------
print("🔄 Descargando archivo banco desde Google Drive...")
banco_stream = read_excel_from_drive(BANCO_FILE_ID)
banco = pd.read_excel(banco_stream, dtype=str, keep_default_na=False)
col_codigo = "CONSE"
if col_codigo not in banco.columns:
    raise ValueError("La columna 'CONSE' no existe en el archivo del banco.")
banco[col_codigo] = banco[col_codigo].apply(norm_code)

# ----------------------------
# LISTAR FICHAS DESDE GOOGLE DRIVE
# ----------------------------
ANIO = "2025"

# ----------------------------
# LISTAR TODAS LAS FICHAS DEL AÑO
# ----------------------------
def list_files_por_anio(folder_id, anio):
    archivos_anio = []
    items = list_files_in_folder(folder_id)
    for f in items:
        if f["mimeType"] == "application/vnd.google-apps.folder":
            # Si la carpeta se llama como el año actual, agregamos sus archivos
            if f["name"] == anio:
                archivos = list_files_in_folder(f["id"])
                for a in archivos:
                    if a["name"].lower().endswith((".xlsx", ".xlsm")):
                        archivos_anio.append(a)
            else:
                # Si es otra carpeta, seguimos buscando recursivamente
                archivos_anio.extend(list_files_por_anio(f["id"], anio))
        else:
            # Ignoramos archivos que no estén dentro de carpeta del año
            continue
    return archivos_anio

print(f"📂 Buscando fichas del año {ANIO} en subcarpetas...")
files_anio = list_files_por_anio(FICHAS_FOLDER_ID, ANIO)
print(f"🔍 Se encontraron {len(files_anio)} fichas para procesar.\n")

registros = []

# ===================================================================
#                     PROCESAMIENTO DE FICHAS DESDE DRIVE
# ===================================================================
for f in files_anio:
    filename = f["name"]
    file_id = f["id"]

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        print(f"⏭ Saltando archivo no Excel: {filename}")
        continue

    try:
        print(f"📥 Leyendo ficha: {filename}")
        stream = read_excel_from_drive(file_id)
        wb = load_workbook(stream, data_only=True)
        nombre_base = filename.lower().replace(".xlsx", "").replace(".xlsm", "")

        # ----------------------------
        # Buscar hoja principal
        # ----------------------------
        posibles_codigos = [p for p in nombre_base.split() if "ind" in p.lower() and "-" in p]
        nombre_hoja = None
        for hoja in wb.sheetnames:
            if any(c in hoja.lower() for c in posibles_codigos):
                nombre_hoja = hoja
                break
            if nombre_base in hoja.lower():
                nombre_hoja = hoja
                break
        if not nombre_hoja:
            nombre_hoja = next(
                (name for name in wb.sheetnames if "ficha" in name.lower() or "indicador" in name.lower()),
                wb.sheetnames[0]
            )

        ws = wb[nombre_hoja]

        # ----------------------------
        # Extraer código (L5:M5)
        # ----------------------------
        codigo = None
        for merged in ws.merged_cells.ranges:
            if "L5" in str(merged) or "M5" in str(merged):
                codigo = ws[merged.coord.split(":")[0]].value
                break
        if codigo is None:
            codigo = ws["L5"].value or ws["M5"].value

        codigo = norm_code(codigo)

        # ----------------------------
        # Validar código con archivo y hoja
        # ----------------------------
        nombre_base_mayus = nombre_base.upper().replace(" ", "")
        hoja_mayus = nombre_hoja.upper().replace(" ", "")
        codigo_valido = codigo and (codigo in nombre_base_mayus or codigo in hoja_mayus)

        if not codigo_valido:
            registros.append({
                "archivo": filename,
                "hoja": nombre_hoja,
                "codigo": codigo,
                "accion": "alerta_inconsistencia",
                "ok": False
            })
            continue

        # ----------------------------
        # Campos hoja principal
        # ----------------------------
        fila = {
            col_codigo: codigo,
            "INDICADOR": clean_str(ws["C5"].value),
            "JERARQUÍA": clean_str(ws["I5"].value),
            "PROCESO": clean_str(ws["H7"].value),
            "OBJETIVO-DESCRIPCIÓN": clean_str(ws["C6"].value),
            "ÁREA": clean_str(ws["C7"].value),
            "TIPO DE INDICADOR": clean_str(ws["C8"].value),
            "TENDENCIA": clean_str(ws["L8"].value),
            "FÓRMULA": f"{clean_str(ws['C9'].value)} / {clean_str(ws['H9'].value)}",
            "FUENTE NUMERADOR": clean_str(ws["C10"].value),
            "FUENTE DENOMINADOR": clean_str(ws["H10"].value),
            "PERIODICIDAD MEDICION": clean_str(ws["C11"].value),
            "PERIODICIDAD ANÁLISIS": clean_str(ws["C12"].value),
            "OBSERVACIONES": clean_str(ws["C13"].value),
            "NORMA RELACIONADA": clean_str(ws["L9"].value),
            "Critico": clean_str(ws["K11"].value),
            "Aceptable": clean_str(ws["L11"].value),
            "Satisfactorio": clean_str(ws["M11"].value)
        }

        # ----------------------------
        # Meses 2025
        # ----------------------------
        mapa_meses = {
            "ene-25": "B19", "feb-25": "C19", "mar-25": "D19", "abr-25": "E19",
            "may-25": "F19", "jun-25": "G19", "jul-25": "H19", "ago-25": "I19",
            "sept-25": "J19", "oct-25": "K19", "nov-25": "L19", "dic-25": "M19"
        }

        for mes, celda in mapa_meses.items():
            valor = ws[celda].value
            if valor is None or str(valor).strip() == "" or (isinstance(valor, str) and "#" in valor):
                fila[mes] = "N/A"
                continue
            try:
                num = float(valor)
                if 0 < num <= 1:
                    fila[mes] = f"{num * 100:.2f}%"
                elif num.is_integer():
                    fila[mes] = str(int(num))
                else:
                    fila[mes] = str(num)
            except:
                fila[mes] = "N/A"

        # ----------------------------
        # Cálculo anual 2025
        # ----------------------------
        columnas_meses = list(mapa_meses.keys())
        valores_numericos = []
        es_porcentaje = False
        for col in columnas_meses:
            val = fila.get(col)
            if val in (None, "N/A"):
                continue
            s = str(val).strip()
            if s.endswith("%"):
                try:
                    valores_numericos.append(float(s.replace("%", "")))
                    es_porcentaje = True
                except:
                    pass
                continue
            try:
                valores_numericos.append(float(s))
            except:
                pass

        if valores_numericos:
            prom = sum(valores_numericos) / len(valores_numericos)
            fila["VALOR ANUAL 2025"] = f"{prom:.2f}%" if es_porcentaje else round(prom, 2)
        else:
            fila["VALOR ANUAL 2025"] = "Error"

        # ----------------------------
        # Hoja de evaluación
        # ----------------------------
        hoja_eval = next((wb[h] for h in wb.sheetnames if "eval" in h.lower()), None)
        if hoja_eval:
            fila_eval = {
                "ESTADO DEL INDICADOR": clean_str(hoja_eval["A2"].value),
                "ORIGEN": clean_str(hoja_eval["B2"].value),
                "DOCUMENTADO": clean_str(hoja_eval["C2"].value),
                "DE SEG CONTRACTUAL": clean_str(hoja_eval["D2"].value),
                "REVISADOS": clean_str(hoja_eval["E2"].value),
                "VALORACIÓN 2025": clean_str(hoja_eval["G2"].value),
            }
            fila.update({k: v for k, v in fila_eval.items() if k in banco.columns})

        # ----------------------------
        # Guardar solo campos existentes
        # ----------------------------
        fila = {k: v for k, v in fila.items() if k in banco.columns}

        # ----------------------------
        # Actualizar o agregar
        # ----------------------------
        if codigo in banco[col_codigo].values:
            for k, v in fila.items():
                banco.loc[banco[col_codigo] == codigo, k] = v
            accion = "actualizado"
        else:
            nueva = {c: None for c in banco.columns}
            nueva.update(fila)
            banco = pd.concat([banco, pd.DataFrame([nueva])], ignore_index=True)
            accion = "agregado"

        # ----------------------------
        # Registro (log)
        # ----------------------------
        registros.append({
            "archivo": filename,
            "hoja": nombre_hoja,
            "codigo": codigo,
            "accion": accion,
            "ok": True
        })

    except Exception as e:
        registros.append({
            "archivo": filename,
            "hoja": None,
            "codigo": None,
            "accion": f"error: {e}",
            "ok": False
        })

# ===================================================================
# GUARDAR BANCO DIRECTAMENTE EN DRIVE CON ESTILOS
# ===================================================================
print("🎨 Aplicando estilos al banco en memoria...")
# Guardamos en BytesIO para aplicar estilos con openpyxl
with io.BytesIO() as buffer:
    banco.to_excel(buffer, index=False)
    buffer.seek(0)
    wb_banco = load_workbook(buffer)
    ws_banco = wb_banco.active

    header_fill = PatternFill(start_color="A7D08C", end_color="A7D08C", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    columnas_sin_color = ["Q", "R", "S"]

    # Encabezados
    for col in range(1, ws_banco.max_column + 1):
        col_letter = get_column_letter(col)
        cell = ws_banco.cell(row=1, column=col)
        if col_letter not in columnas_sin_color:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    # Colores especiales
    ws_banco["Q1"].fill = PatternFill(start_color="FF0000", fill_type="solid")
    ws_banco["R1"].fill = PatternFill(start_color="FFFF00", fill_type="solid")
    ws_banco["S1"].fill = PatternFill(start_color="00FF00", fill_type="solid")

    # Bordes y alineación en todas las celdas
    for row in ws_banco.iter_rows(min_row=2, max_row=ws_banco.max_row, max_col=ws_banco.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Ajuste de columnas
    for col in ws_banco.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_banco.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Ajuste de filas
    for row in range(2, ws_banco.max_row + 1):
        ws_banco.row_dimensions[row].height = 30

    # Filtro
    ws_banco.auto_filter.ref = f"A1:{get_column_letter(ws_banco.max_column)}{ws_banco.max_row}"

    # Guardar de nuevo en buffer
    buffer_out = io.BytesIO()
    wb_banco.save(buffer_out)
    buffer_out.seek(0)
    # Subir a Drive
    upload_bytes_to_drive(buffer_out.read(), BANCO_FILE_ID)

print("✔ Banco actualizado y con estilos aplicado directamente en Drive ✅")

# ===================================================================
# REPORTE FINAL EN MEMORIA
# ===================================================================
df_rep = pd.DataFrame(registros)
print("📄 Reporte final en memoria listo.")

# ===================================================================
# REPORTE FINAL EN MEMORIA Y SUBIDA A DRIVE
# ===================================================================

# ID de los archivos existentes en Drive
FILE_ID_REPORTE_EXCEL = "10OFq9GldfDwY3MrH57WnwJY36Lkcv_gD"
FILE_ID_REPORTE_CSV   = "1Yeksk5u5dpWkn8EaCeGWuJ8qqJNfB7Ln"

# Subir reporte Excel a Drive
with io.BytesIO() as buffer_excel:
    with pd.ExcelWriter(buffer_excel, engine="openpyxl") as writer:
        df_rep.to_excel(writer, sheet_name="Reporte Completo", index=False)
        df_rep[df_rep["accion"] == "actualizado"].to_excel(writer, sheet_name="Actualizados", index=False)
        df_rep[df_rep["accion"] == "agregado"].to_excel(writer, sheet_name="Agregados", index=False)
        df_rep[df_rep["ok"] == False].to_excel(writer, sheet_name="Errores", index=False)
    buffer_excel.seek(0)
    upload_bytes_to_drive(buffer_excel.read(), FILE_ID_REPORTE_EXCEL)

# Subir reporte CSV a Drive
with io.BytesIO() as buffer_csv:
    df_rep.to_csv(buffer_csv, index=False)
    buffer_csv.seek(0)
    upload_bytes_to_drive(buffer_csv.read(), FILE_ID_REPORTE_CSV)

print("📄 Reporte final actualizado correctamente en Drive ✅")
